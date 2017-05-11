# -*- coding: utf-8 -*-
"""
Created on Fri May  5 09:28:32 2017

@author: andres
"""
from __future__ import absolute_import, division, print_function, unicode_literals
import Tkinter
from Tkinter import Toplevel
import pandas as pd
import openpyxl
import locale

locale.setlocale(locale.LC_ALL, '')
from Tkinter import Message, Button, END, Text
import ystockquote
from pandas.tseries.offsets import BDay
from pandas.tseries.holiday import USFederalHolidayCalendar as calendar
import time
# from datetime import datetime, timedelta
######
import datetime as dt
from pandas.tseries.holiday import AbstractHolidayCalendar, Holiday, nearest_workday, \
	USMartinLutherKingJr, USPresidentsDay, GoodFriday, USMemorialDay, \
	USLaborDay, USThanksgivingDay

import threading


class USTradingCalendar(AbstractHolidayCalendar):
	rules = [Holiday('NewYearsDay', month=1, day=1, observance=nearest_workday),
			 USMartinLutherKingJr,
			 USPresidentsDay,
			 GoodFriday,
			 USMemorialDay,
			 Holiday('USIndependenceDay', month=7, day=4, observance=nearest_workday),
			 USLaborDay,
			 USThanksgivingDay,
			 Holiday('Christmas', month=12, day=25, observance=nearest_workday)
			 ]


def get_trading_close_holidays(year):
	inst = USTradingCalendar()
	return inst.holidays(dt.datetime(year - 1, 12, 31), dt.datetime(year, 12, 31))


#####
interval = 5 # in seconds
compData = {}
companyData = {}



def format_volume_string(vol):
	vol_str = locale.format("%d", vol, grouping=True)
	if vol>=1000000 and vol<1000000000:
		vol_str = str(round(vol/1000000.0,2))+'M'
	elif vol>=1000000000:
		vol_str = str(round(vol/1000000000.0,2))+'B'
	return vol_str



class simpleapp_tk(Tkinter.Tk):
	def __init__(self, parent):
		Tkinter.Tk.__init__(self, parent)
		self.parent = parent
		self.initialize()

	def initialize(self):
		self.update_threadRunning = True

		self.grid()
		self.geometry("1000x900")

		canvas = Tkinter.Canvas(self, height=200)
		canvas.config(background="black")
		frame = Tkinter.Frame(canvas)
		frame.configure(background="black")
		scrollbar = Tkinter.Scrollbar(self, orient="vertical", command=canvas.yview)

		canvas.configure(yscrollcommand=scrollbar.set)
		scrollbar.pack(side="right", fill="y")
		canvas.pack(side="left", fill="both", expand=True)
		canvas.create_window((4, 4), window=frame, anchor="nw", tags="frame")
		frame.bind("<Configure>", lambda x: canvas.configure(scrollregion=canvas.bbox("all")))
		self.bind("<Down>", lambda x: canvas.yview_scroll(3, 'units'))  # bind "Down" to scroll down
		self.bind("<Up>", lambda x: canvas.yview_scroll(-3, 'units'))  # bind "Up" to scroll up
		# bind the mousewheel to scroll up/down
		self.bind("<MouseWheel>", lambda x: canvas.yview_scroll(int(-1 * (x.delta / 40)), "units"))


		self.labelVariable = Tkinter.StringVar()
		label = Tkinter.Label(frame, textvariable=self.labelVariable, anchor="w", fg="gold", bg="black")
		label.grid(column=0, row=1, columnspan=9, sticky='EW')
		self.labelVariable.set(
			u" Ticker:\tPrice:\tChng:\t% Chng:\t\t1 dy:\t1 wk:\t1 mo:\t3 mo:\t\tVol:\t90d Avg Vol:")

		self.childTickerWindows = {}
		# self.allStocksList = set(companyData.keys()+compData.keys())
		self.allStocksList = set(companyData)
		self.companyData_StringVar = {}
		i = 2
		print(len(companyData))
		for item in companyData:
			self.companyData_StringVar[item] = Tkinter.StringVar()
			label = Tkinter.Label(frame, textvariable=self.companyData_StringVar[item], anchor="w", fg="gold",
								  bg="black")
			label.grid(column=0, row=i, columnspan=9, sticky='EW')
			#label.pack()

			self.companyData_StringVar[item].set(
				'  {:<5}\t{:>8.2f}\t{:>6.2f}\t{:>8.2f}\t\t{:>6.2f}\t{:>6.2f}\t{:>6.2f}\t{:>6.2f}\t{:>16}\t{:>16}\t'.format(
					item, companyData[item]['last'], companyData[item]['change'], companyData[item]['pctChange'],
					companyData[item]['ret1'], companyData[item]['ret7'], companyData[item]['ret30'],
					companyData[item]['ret90'], format_volume_string(companyData[item]['vol']), format_volume_string(companyData[item]['volAvg'])))

			button = Tkinter.Button(frame, text=item, command=lambda r=item: self.OnButtonClick(r))
			button.config(fg="gold", bg="black")
			button.grid(column=11, row=i)
			i += 1

		self.grid_columnconfigure(0, weight=1)
		self.resizable(False, True)
		self.update()
		self.geometry(self.geometry())

	def OnButtonClick(self, button_id):
		if button_id in self.childTickerWindows:
			if self.childTickerWindows[button_id].winfo_exists():  # window exists; close and re-open
				self.childTickerWindows[button_id].destroy()

		top = Toplevel()
		top.title('Comps for ' + str(button_id))
		top.geometry('900x300-30+30')
		top.config(bg="black")

		self.childTickerWindows[button_id] = top

		T = Text(top, height=15, width=115, background='black', foreground='gold')
		T.pack()
		# T.tag_config(background="gray", foreground="gold")
		T.insert(END,
				 "Ticker:\t    Price:\t   Chng:\t  % Chng:\t   1 dy: \t  1wk: \t 1 mo:\t\t   3 mo:\t\t Vol:\t\t90d Avg Vol:\n\n")


		T.insert(END,
				 '  {:<5}\t{:>10.2f}\t{:>8.2f}\t{:>8.2f}\t\t{:>6.2f}\t{:>6.2f}\t{:>6.2f}\t  {:>6.2f}\t  {:>16}\t{:>16}\n\n'.format(
					 str(button_id), companyData[button_id]['last'], companyData[button_id]['change'],
					 companyData[button_id]['pctChange'], companyData[button_id]['ret1'],
					 companyData[button_id]['ret7'], companyData[button_id]['ret30'], companyData[button_id]['ret90'],
					 format_volume_string(companyData[button_id]['vol']), format_volume_string(companyData[button_id]['volAvg'])))



		for comp_counter in range(1, 9+1):
			comp_str = 'comp'+str(comp_counter)
			if comp_str in companyData[button_id].keys() and companyData[button_id][comp_str] != '':





				T.insert(END,
						 '  {:<5}\t{:>10.2f}\t{:>8.2f}\t{:>8.2f}\t\t{:>6.2f}\t{:>6.2f}\t{:>6.2f}\t  {:>6.2f}\t  {:>16}\t{:>16}\n'.format(
							 str(companyData[button_id][comp_str]), compData[companyData[button_id][comp_str]]['last'],
							 compData[companyData[button_id][comp_str]]['change'],
							 compData[companyData[button_id][comp_str]]['pctChange'],
							 compData[companyData[button_id][comp_str]]['ret1'],
							 compData[companyData[button_id][comp_str]]['ret7'],
							 compData[companyData[button_id][comp_str]]['ret30'],
							 compData[companyData[button_id][comp_str]]['ret90'],
							 format_volume_string(compData[companyData[button_id][comp_str]]['vol']),
							 format_volume_string(compData[companyData[button_id][comp_str]]['volAvg'])))


		T.pack()

		# Text.insert(END, "Ticker:\tPrice:\tChng:\t% Chng:\t1 dy:\t1 wk:\t1 mo:\t3 mo:\tVol:\t90d Avg Vol:")
		# msg.pack()

		button = Button(top, text='Close', command=top.destroy)
		button.pack()

	def OnChildClose(self):
		self.wButton.config(state='normal')
		self.top.destroy()

	def OnPressEnter(self, event):
		self.labelVariable.set(self.entryVariable.get() + " (You pressed enter)")
		self.entry.focus_set()
		self.entry.selection.range(0, Tkinter.END)


def update_data(mainobject):
	print('Update starting')
	# Checks for change in price in stocks that are in companyData....
	for item in mainobject.allStocksList:
		print('Updating %s' % item)
		t = 0
		maximum_errors_reached = False
		while (True):
			try:
				last = float(ystockquote.get_last_trade_price(item))
				vol = int(ystockquote.get_volume(item))

				break
			except:
				t += 1
			if (t == 3):
				maximum_errors_reached = True
				break
		if maximum_errors_reached:
			continue


		vol_str = format_volume_string(vol)

		if item in companyData:
			companyData[item]['last'] = last
			companyData[item]['change'] = round(last - companyData[item]['prevClose'], 2)
			companyData[item]['pctChange'] = round(companyData[item]['change'] / companyData[item]['prevClose'] * 100,  2)
			companyData[item]['vol'] = vol

			avg_vol_str = format_volume_string(companyData[item]['volAvg'])

			mainobject.companyData_StringVar[item].set('  {:<5}\t{:>8.2f}\t{:>6.2f}\t{:>8.2f}\t\t{:>6.2f}\t{:>6.2f}\t{:>6.2f}\t{:>6.2f}\t{:>16}\t{:>16}\t'.format(item, companyData[item]['last'], companyData[item]['change'], companyData[item]['pctChange'], companyData[item]['ret1'], companyData[item]['ret7'], companyData[item]['ret30'], companyData[item]['ret90'], vol_str, avg_vol_str))


		if item in compData:
			compData[item]['last'] = last
			compData[item]['change'] = round(last - compData[item]['prevClose'], 2)
			compData[item]['pctChange'] = round(compData[item]['change'] / compData[item]['prevClose'] * 100, 2)
			compData[item]['vol'] = vol

	print('Updated')


def update_data_threadfunction(mainobject):
	while mainobject.update_threadRunning:
		update_data(mainobject)
		time.sleep(interval)


####################################
# organize calendar
today = pd.datetime.today()
today = today.replace(hour=0, minute=0, second=0, microsecond=0)
year = dt.date.today().year

oneDy = today - BDay(2)
oneDy = oneDy.to_pydatetime()

oneWk = today - BDay(6)
oneWk = oneWk.to_pydatetime()

oneMo = today - BDay(23)
oneMo = oneMo.to_pydatetime()

threeMo = today - BDay(91)
threeMo = threeMo.to_pydatetime()

# check for holidays
# dr = pd.date_range(start=threeMo, end=today) # note that this will include all calendar days, not just business
# cal = calendar()
# holidays = cal.holidays(start=threeMo, end = today).to_pydatetime()
holidays = get_trading_close_holidays(year)

while today in holidays:  # today
	today -= BDay(1)
	today = today.to_pydatetime()
	oneDy = today - BDay(1)
	oneDy = oneDy.to_pydatetime()

while oneDy in holidays:  # yesterday
	oneDy -= BDay(1)
	oneDy = oneDy.to_pydatetime()

while oneWk in holidays:  # one week
	oneWk -= BDay(1)
	oneWk = oneWk.to_pydatetime()

while oneMo in holidays:  # one month
	oneMo -= BDay(1)
	oneMo = oneMo.to_pydatetime()

while threeMo in holidays:  # three months
	threeMo += BDay(1)
	threeMo = threeMo.to_pydatetime()

# convert datetime object to string, for API query
today = today.strftime('%Y-%m-%d')
oneDy = oneDy.strftime('%Y-%m-%d')
oneWk = oneWk.strftime('%Y-%m-%d')
oneMo = oneMo.strftime('%Y-%m-%d')
threeMo = threeMo.strftime('%Y-%m-%d')


####################################

# load data from xlsx and parse them

filename = 'Spectrum Comps Lg Cap.xlsx'
print('Reading portfolio file...')
wb = openpyxl.load_workbook(filename)
sheet = wb.get_sheet_by_name('Comps')

for row in range(2, sheet.max_row + 1):
	if sheet['A' + str(row)].value == 'x':
		ticker = str(sheet['B' + str(row)].value)
		companyName = str(sheet['C' + str(row)].value)
		comp1 = sheet['E' + str(row)].value
		comp2 = sheet['F' + str(row)].value
		comp3 = sheet['G' + str(row)].value
		comp4 = sheet['H' + str(row)].value
		comp5 = sheet['I' + str(row)].value
		comp6 = sheet['J' + str(row)].value
		comp7 = sheet['K' + str(row)].value
		comp8 = sheet['L' + str(row)].value
		comp9 = sheet['M' + str(row)].value

		companyData.setdefault(ticker, {'companyName': '', 'last': 0, 'prevClose': 0, 'change': 0,
										'pctChange': 0, 'ret1': 0, 'ret7': 0,
										'ret30': 0, 'ret90': 0, 'vol': 0,
										'volAvg': 0, 'volDelta': 0, 'comp1': '',
										'comp2': '', 'comp3': '', 'comp4': '',
										'comp5': '', 'comp6': '', 'comp7': '',
										'comp8': '', 'comp9': ''})

		companyData[ticker]['companyName'] = str(companyName)
		if comp1 is not None:
			companyData[ticker]['comp1'] = str(comp1).strip()
		if comp2 is not None:
			companyData[ticker]['comp2'] = str(comp2).strip()
		if comp3 is not None:
			companyData[ticker]['comp3'] = str(comp3).strip()
		if comp4 is not None:
			companyData[ticker]['comp4'] = str(comp4).strip()
		if comp5 is not None:
			companyData[ticker]['comp5'] = str(comp5).strip()
		if comp6 is not None:
			companyData[ticker]['comp6'] = str(comp6).strip()
		if comp7 is not None:
			companyData[ticker]['comp7'] = str(comp7).strip()
		if comp8 is not None:
			companyData[ticker]['comp8'] = str(comp8).strip()
		if comp9 is not None:
			companyData[ticker]['comp9'] = str(comp9).strip()

print('...........................done')

# filename = 'Spectrum Metrics Factset.xlsx'
print('Preparing comps and metrics...')

# wb = openpyxl.load_workbook(filename, data_only=True)
# sheet = wb.get_sheet_by_name('Sheet1')

compList = []
for key in companyData:
	if key not in compList:
		compList.append(key)
	if companyData[key]['comp1'] is not None and companyData[key]['comp1'] != '' and companyData[key][
		'comp1'] not in compList:
		compList.append(companyData[key]['comp1'])
	if companyData[key]['comp2'] is not None and companyData[key]['comp2'] != '' and companyData[key][
		'comp2'] not in compList:
		compList.append(companyData[key]['comp2'])
	if companyData[key]['comp3'] is not None and companyData[key]['comp3'] != '' and companyData[key][
		'comp3'] not in compList:
		compList.append(companyData[key]['comp3'])
	if companyData[key]['comp4'] is not None and companyData[key]['comp4'] != '' and companyData[key][
		'comp4'] not in compList:
		compList.append(companyData[key]['comp4'])
	if companyData[key]['comp5'] is not None and companyData[key]['comp5'] != '' and companyData[key][
		'comp5'] not in compList:
		compList.append(companyData[key]['comp5'])
	if companyData[key]['comp6'] is not None and companyData[key]['comp6'] != '' and companyData[key][
		'comp6'] not in compList:
		compList.append(companyData[key]['comp6'])
	if companyData[key]['comp7'] is not None and companyData[key]['comp7'] != '' and companyData[key][
		'comp7'] not in compList:
		compList.append(companyData[key]['comp7'])
	if companyData[key]['comp8'] is not None and companyData[key]['comp8'] != '' and companyData[key][
		'comp8'] not in compList:
		compList.append(companyData[key]['comp8'])
	if companyData[key]['comp9'] is not None and companyData[key]['comp9'] != '' and companyData[key][
		'comp9'] not in compList:
		compList.append(companyData[key]['comp9'])


for item in compList:
	ticker = item

	print(ticker)

	last = ystockquote.get_last_trade_price(ticker)
	if last=='N/A':
		print('Ticker ',ticker,' returned NO LAST PRICE')
		last = '1'
	last = float(last)

	prevClose = ystockquote.get_previous_close(ticker)
	if prevClose=='N/A':
		print('Ticker ',ticker,' returned NO PREV CLOSE PRICE')
		prevClose = '1'
	prevClose = float(prevClose)

	change = round(last - prevClose, 2)
	pctChange = round(change / prevClose * 100, 2)

	vol = ystockquote.get_volume(ticker)
	if vol=='N/A':
		print('Ticker ',ticker,' returned NO VOLUME')
		vol = '0'
	vol = int(vol)

	volAvg = ystockquote.get_average_daily_volume(ticker)
	if volAvg=='N/A':
		print('Ticker ',ticker,' returned NO AVG VOLUME')
		volAvg = '0'
	volAvg = int(volAvg)


	try:
		oneDy_price = float(ystockquote.get_historical_prices(ticker, oneDy, oneDy )[oneDy]['Adj Close'])
	except:
		oneDy_price = last

	try:
		oneWk_price = float(ystockquote.get_historical_prices(ticker, oneWk, oneWk )[oneWk]['Adj Close'])
	except:
		oneWk_price = last

	try:
		oneMo_price = float(ystockquote.get_historical_prices(ticker, oneMo, oneMo )[oneMo]['Adj Close'])
	except:
		oneMo_price = last

	try:
		threeMo_price = float(ystockquote.get_historical_prices(ticker, threeMo, threeMo )[threeMo]['Adj Close'])
	except:
		threeMo_price = last

	ret1 = round((prevClose / oneDy_price - 1) * 100, 2)
	ret7 = round((prevClose / oneWk_price - 1) * 100, 2)
	ret30 = round((prevClose / oneMo_price - 1) * 100, 2)
	ret90 = round((prevClose / threeMo_price - 1) * 100, 2)



	#### Data is now collected.  If in companyData, fill in information, in addition fill in compData

	### populate info from companyData, revise above to do so.....
	if ticker in companyData:
		companyData[ticker]['ret1'] = ret1
		companyData[ticker]['ret7'] = ret7
		companyData[ticker]['ret30'] = ret30
		companyData[ticker]['ret90'] = ret90
		companyData[ticker]['vol'] = vol
		companyData[ticker]['volAvg'] = volAvg
		# omitting voldDelta for now, I don't think it's useful
		# companyData[ticker]['volDelta'] = (companyData[ticker]['vol'] / companyData[ticker]['volAvg'] - 1) * 100
		companyData[ticker]['last'] = round(last, 2)
		companyData[ticker]['change'] = change
		companyData[ticker]['pctChange'] = pctChange
		companyData[ticker]['prevClose'] = prevClose

	compData.setdefault(ticker, {'companyName': '', 'prevClose': 0, 'ret1': 0, 'ret7': 0,
								 'ret30': 0, 'ret90': 0, 'last': 0,
								 'change': 0, 'pctChange': 0,
								 'vol': 0, 'volAvg': 0, 'volDelta': 0})
	compData[ticker]['companyName'] = companyName
	compData[ticker]['ret1'] = ret1
	compData[ticker]['ret7'] = ret7
	compData[ticker]['ret30'] = ret30
	compData[ticker]['ret90'] = ret90
	compData[ticker]['vol'] = vol
	compData[ticker]['volAvg'] = volAvg
	# compData[ticker]['volDelta'] = (compData[ticker]['vol'] / compData[ticker]['volAvg'] - 1) * 100
	compData[ticker]['last'] = round(last, 2)
	compData[ticker]['change'] = change
	compData[ticker]['pctChange'] = pctChange
	compData[ticker]['prevClose'] = prevClose


print('...........................done')

def on_closing():
	global app

	print("APP CLOSING")

	app.update_threadRunning = False

	time.sleep(0.1)
	app.destroy()

app = simpleapp_tk(None)
app.title('Spectrum Metrics')

app.protocol("WM_DELETE_WINDOW", on_closing)

update_data_thread = threading.Thread(target=update_data_threadfunction, args=[app])
update_data_thread.daemon = True
update_data_thread.start()

app.mainloop()

